from Db import *
from Cafe import *
from Owner import *
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox
import design
import sys
import os
import hashlib
import openpyxl as ex
import uuid
import pandas as pd
import pdfkit

db = Db()
strnumber = 1


class ExampleApp(QtWidgets.QMainWindow, design.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.comboBox.activated[str].connect(self.Choise)
        self.SearchButton.clicked.connect(self.Search)
        self.UpdateButton.clicked.connect(self.Save)
        self.RightButton.clicked.connect(self.Right)
        self.LeftButton.clicked.connect(self.Left)
        self.DelButton.clicked.connect(self.Delete)
        self.SyncButton.clicked.connect(self.SyncDB)
        self.ExButton.clicked.connect(self.ToExcel)
        self.PDFButton.clicked.connect(self.ToPdf)

    def Right(self):
        global strnumber
        strnumber += 1
        try:
            self.Show(number=strnumber)
        except:
            QMessageBox.about(self, "Error!!!", "Дальше листать нельзя")
            strnumber -= 1
        self.Label.setText(str(strnumber))

    def ToArray(self, arr):
        result = []
        if self.comboBox.currentText() == 'Заведение':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j == 3:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Готовка':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j == 2:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(str(arr[i][j]))
                result.append(buf)

        if self.comboBox.currentText() == 'Блюдо':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 1:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Ингридиент':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 1:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Владелец':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Продукт':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 2:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Поставщик':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Рецепт':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Столик':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 1:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Гость':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 1:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Официант':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    buf.append(arr[i][j])
                result.append(buf)

        if self.comboBox.currentText() == 'Повар':
            for i in range(len(arr)):
                buf = []
                for j in range(len(arr[i])):
                    if j > 2:
                        buf.append(arr[i][j][0])
                    else:
                        buf.append(arr[i][j])
                result.append(buf)

        return result

    def ToPdf(self):
        new = ex.Workbook()
        filename = uuid.uuid4().hex
        try:
            data = ex.load_workbook(filename=filename, read_only=False)
        except:
            filename = uuid.uuid4().hex
        new_sheet = new.active
        result = db.Select(table=self.comboBox.currentText(), strnumber=strnumber)
        result = self.ToArray(result)
        for i in range(len(result)):
            for j in range(len(result[i])):
                new_sheet.cell(i + 1, j + 1).value = result[i][j]
        new.save(filename+".xlsx")
        new.close()
        print(filename)
        df = pd.read_excel(filename+".xlsx")
        df.to_html(filename+".html")
        pdfkit.from_file(filename+".html", filename+".pdf")
        #os.system('start excel.exe ' + filename)

    def ToExcel(self):
        new = ex.Workbook()
        filename = uuid.uuid4().hex
        try:
            data = ex.load_workbook(filename=filename, read_only=False)
        except:
            filename = uuid.uuid4().hex
        filename += ".xlsx"
        new_sheet = new.active
        result = db.Select(table=self.comboBox.currentText(), strnumber=strnumber)
        result = self.ToArray(result)
        for i in range(len(result)):
            for j in range(len(result[i])):
                new_sheet.cell(i + 1, j + 1).value = result[i][j]
        new.save(filename)
        new.close()
        os.system('start excel.exe ' + filename)

    def Left(self):
        global strnumber
        if strnumber > 1:
            strnumber -= 1
            self.Show(number=strnumber)
            self.Label.setText(str(strnumber))

    def md5(self, fname):
        hash_md5 = hashlib.md5()
        with open(fname, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def exec_sql_file(self, cursor, sql_file):
        statement = ""
        for line in open(sql_file, encoding='utf-8'):
            if re.match(r'--', line):
                continue
            if not re.search(r';$', line):
                statement = statement + line
            else:
                statement = statement + line
                try:
                    cursor.execute(statement)
                except:
                    print("MySQLError during execute statement")
                statement = ""

    def CreateDb(self):
        directory = os.getcwd()
        files = os.listdir(directory)
        arr = []
        for i in files:
            if i[-4:] == '.sql':
                if len(re.sub('[^0-9]+', '', i)) != 4:
                    QMessageBox.about(self, "Error", "Invalid .sql file name")
                    return
                arr.append(i)
        for i in arr:
            cursor = db.conn.cursor()
            f = open(i)
            cursor.execute(f.read())
            cursor.close()

    def SyncDB(self):
        directory = os.getcwd()
        files = os.listdir(directory)
        arr = []
        for i in files:
            if i[-4:] == '.sql':
                if len(re.sub('[^0-9]+', '', i)) != 4:
                    QMessageBox.about(self, "Error", "Invalid .sql file name")
                    return
                arr.append(i)
        for i in arr:
            cursor = db.conn.cursor()
            try:
                cursor.execute("SELECT * FROM cswl3orm.files WHERE filename = '%s'" % i)
            except:
                cursor.close()
                self.CreateDb()
                return
            res = cursor.fetchall()
            cursor.close()
            if len(res) == 0:
                cursor = db.conn.cursor()
                try:
                    self.exec_sql_file(cursor, i)
                except:
                    QMessageBox.about(self, "Error!", "Invalid SqlFile %s" % i)
                cursor.close()
                cursor = db.conn.cursor()
                cursor.execute("INSERT INTO cswl3orm.files(FileName, FileHash)VALUES('%s', '%s')" % (i, self.md5(i)))
                cursor.execute("COMMIT")
                cursor.close()
            else:
                if self.md5(i) == res[0][1]:
                    QMessageBox.about(self, "Message", "You are using the current version of the database")
                    return
                else:
                    QMessageBox.about(self, "Error!", "The content of the file has been changed %s" % i)
                    return
        QMessageBox.about(self, "Message!", "Sync of Database ended")

    def Search(self):
        self.Show(subres=db.Search(self.comboBox.currentText(), self.SearchText.text(), strnumber))

    def Show(self, number=1, subres=None):
        self.tableWidget.setRowCount(0)
        if self.comboBox.currentText() == 'Повар':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 3, QtWidgets.QTableWidgetItem(str(result[i][2])))
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][3])
                self.tableWidget.setCellWidget(i, 2, combobox)
            buf = Cafe.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 2, combobox)

        if self.comboBox.currentText() == 'Заведение':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][2])))
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][3])
                self.tableWidget.setCellWidget(i, 3, combobox)
            buf = Owner.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 3, combobox)

        if self.comboBox.currentText() == 'Готовка':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][1])))
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][2])
                self.tableWidget.setCellWidget(i, 1, combobox)
            buf = Cook.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 1, combobox)

        if self.comboBox.currentText() == 'Блюдо':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                # Recipe
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][2])
                self.tableWidget.setCellWidget(i, 2, combobox)
                # Creating
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][3])
                self.tableWidget.setCellWidget(i, 3, combobox)
                # Visitor
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][4])
                self.tableWidget.setCellWidget(i, 4, combobox)
            # Recipe
            buf = Recipe.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 2, combobox)
            # Creating
            buf = Creating.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + str(buf[i].Time))
            self.tableWidget.setCellWidget(count - 1, 3, combobox)
            # Visitor
            buf = Visitor.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 4, combobox)

        if self.comboBox.currentText() == 'Ингридиент':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                # Recipe
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][2])
                self.tableWidget.setCellWidget(i, 2, combobox)
                # Product
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][3])
                self.tableWidget.setCellWidget(i, 3, combobox)
            # Recipe
            buf = Recipe.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 2, combobox)
            # Product
            buf = Product.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 3, combobox)

        if self.comboBox.currentText() == 'Владелец':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][2])))

        if self.comboBox.currentText() == 'Продукт':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 3, QtWidgets.QTableWidgetItem(str(result[i][2])))
                # Provider
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][3])
                self.tableWidget.setCellWidget(i, 2, combobox)
                # Cafe
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][4])
                self.tableWidget.setCellWidget(i, 4, combobox)
            # Provider
            buf = Provider.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 2, combobox)
            # Cafe
            buf = Cafe.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 4, combobox)

        if self.comboBox.currentText() == 'Поставщик':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][2])))

        if self.comboBox.currentText() == 'Рецепт':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))

        if self.comboBox.currentText() == 'Столик':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][2])
                self.tableWidget.setCellWidget(i, 2, combobox)
            buf = Waiter.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].Name)
            self.tableWidget.setCellWidget(count - 1, 2, combobox)

        if self.comboBox.currentText() == 'Гость':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][1])))
                combobox = QtWidgets.QComboBox()
                combobox.addItems(result[i][2])
                self.tableWidget.setCellWidget(i, 1, combobox)
            buf = Stolik.select()
            combobox = QtWidgets.QComboBox()
            for i in range(len(buf)):
                combobox.addItem(str(buf[i].Id) + " " + buf[i].TableLevel)
            self.tableWidget.setCellWidget(count - 1, 1, combobox)

        if self.comboBox.currentText() == 'Официант':
            if subres is None:
                result = db.Select(table=self.comboBox.currentText(), strnumber=number)
            else:
                result = subres
            count = len(result) + 1
            self.tableWidget.setRowCount(count)
            for i in range(len(result)):
                self.tableWidget.setItem(i, 0, QtWidgets.QTableWidgetItem(str(result[i][0])))
                self.tableWidget.setItem(i, 1, QtWidgets.QTableWidgetItem(str(result[i][1])))
                self.tableWidget.setItem(i, 2, QtWidgets.QTableWidgetItem(str(result[i][2])))

        self.tableWidget.resizeColumnsToContents()

    def Choise(self, text):
        self.SearchText.setText("")
        global strnumber
        strnumber = 1
        self.Label.setText(str(strnumber))
        if text == 'Поставщик':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Наименование  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Номер телефона  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Продукт':
            self.tableWidget.setColumnCount(5)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Наименование  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Поставщик  "))
            self.tableWidget.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem(
                "  Количество  "))
            self.tableWidget.setHorizontalHeaderItem(4, QtWidgets.QTableWidgetItem(
                "  Заведение  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Владелец':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  ФИО  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Номер телефона  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Заведение':
            self.tableWidget.setColumnCount(4)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Название  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Рейтинг  "))
            self.tableWidget.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem(
                "  Владелец  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Повар':
            self.tableWidget.setColumnCount(4)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  ФИО  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Место работы  "))
            self.tableWidget.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem(
                "  Разряд  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()

        if text == 'Готовка':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Повар  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Дата и время  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Блюдо':
            self.tableWidget.setColumnCount(5)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Название  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Рецепт  "))
            self.tableWidget.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem(
                "  Время приготовления  "))
            self.tableWidget.setHorizontalHeaderItem(4, QtWidgets.QTableWidgetItem(
                "  Гость  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Рецепт':
            self.tableWidget.setColumnCount(2)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("  Номер(Id)  "))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Название  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Ингридиент':
            self.tableWidget.setColumnCount(4)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("Номер(Id)"))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Количество  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Рецепт                                                   "))
            self.tableWidget.setHorizontalHeaderItem(3, QtWidgets.QTableWidgetItem(
                "  Продукт                                                  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Гость':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("Номер(Id)"))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Номер столика  "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  ФИО                                                                 "
                "                                                         "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Столик':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("Номер(Id)"))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  Уровень   "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Официант                                                        "
                "                                                                  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        if text == 'Официант':
            self.tableWidget.setColumnCount(3)
            self.tableWidget.setHorizontalHeaderItem(0, QtWidgets.QTableWidgetItem("Номер(Id)"))
            self.tableWidget.setHorizontalHeaderItem(1, QtWidgets.QTableWidgetItem(
                "  ФИО                                                                    "
                "                                                                         "))
            self.tableWidget.setHorizontalHeaderItem(2, QtWidgets.QTableWidgetItem(
                "  Стаж  "))
            self.tableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
            self.tableWidget.resizeColumnsToContents()
        self.Show()

    def Save(self):
        if self.comboBox.currentText() == 'Повар':
            for i in range(self.tableWidget.rowCount()):
                cafe = str(self.tableWidget.cellWidget(i, 2).currentText())
                cafe = cafe[0:cafe.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j != 2:
                        atr.append(self.tableWidget.item(i, j))
                    else:
                        atr.append(cafe)
                try:
                    db.Merge('Повар', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Заведение':
            for i in range(self.tableWidget.rowCount()):
                owner = str(self.tableWidget.cellWidget(i, 3).currentText())
                owner = owner[0:owner.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j != 3:
                        atr.append(self.tableWidget.item(i, j))
                    else:
                        atr.append(owner)
                try:
                    db.Merge('Заведение', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Готовка':
            for i in range(self.tableWidget.rowCount()):
                cook = str(self.tableWidget.cellWidget(i, 1).currentText())
                cook = cook[0:cook.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j != 1:
                        atr.append(self.tableWidget.item(i, j))
                    else:
                        atr.append(cook)
                try:
                    db.Merge('Готовка', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Блюдо':
            for i in range(self.tableWidget.rowCount()):
                recipe = str(self.tableWidget.cellWidget(i, 2).currentText())
                recipe = recipe[0:recipe.index(" ")]
                creating = str(self.tableWidget.cellWidget(i, 3).currentText())
                creating = creating[0:creating.index(" ")]
                visitor = str(self.tableWidget.cellWidget(i, 4).currentText())
                visitor = visitor[0:visitor.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j == 2:
                        atr.append(recipe)
                    elif j == 3:
                        atr.append(creating)
                    elif j == 4:
                        atr.append(visitor)
                    else:
                        atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Блюдо', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Ингридиент':
            for i in range(self.tableWidget.rowCount()):
                recipe = str(self.tableWidget.cellWidget(i, 2).currentText())
                recipe = recipe[0:recipe.index(" ")]
                product = str(self.tableWidget.cellWidget(i, 3).currentText())
                product = product[0:product.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j == 2:
                        atr.append(recipe)
                    elif j == 3:
                        atr.append(product)
                    else:
                        atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Ингридиент', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        if self.comboBox.currentText() == 'Владелец':
            for i in range(self.tableWidget.rowCount()):
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Владелец', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        if self.comboBox.currentText() == 'Продукт':
            for i in range(self.tableWidget.rowCount()):
                provider = str(self.tableWidget.cellWidget(i, 2).currentText())
                provider = provider[0:provider.index(" ")]
                cafe = str(self.tableWidget.cellWidget(i, 4).currentText())
                cafe = cafe[0:cafe.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j == 2:
                        atr.append(provider)
                    elif j == 4:
                        atr.append(cafe)
                    else:
                        atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Продукт', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        if self.comboBox.currentText() == 'Поставщик':
            for i in range(self.tableWidget.rowCount()):
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Поставщик', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        if self.comboBox.currentText() == 'Рецепт':
            for i in range(self.tableWidget.rowCount()):
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Рецепт', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        if self.comboBox.currentText() == 'Столик':
            for i in range(self.tableWidget.rowCount()):
                waiter = str(self.tableWidget.cellWidget(i, 2).currentText())
                waiter = waiter[0:waiter.index(" ")]
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j != 2:
                        atr.append(self.tableWidget.item(i, j))
                    else:
                        atr.append(waiter)
                try:
                    db.Merge('Столик', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Гость':
            for i in range(self.tableWidget.rowCount()):
                table = str(self.tableWidget.cellWidget(i, 1).currentText())
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    if j == 1:
                        atr.append(table)
                    else:
                        atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Гость', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")

        if self.comboBox.currentText() == 'Официант':
            for i in range(self.tableWidget.rowCount()):
                atr = []
                for j in range(self.tableWidget.columnCount()):
                    atr.append(self.tableWidget.item(i, j))
                try:
                    db.Merge('Официант', atr)
                except:
                    QMessageBox.about(self, "Error!!!", "Invalid Value!")
        self.Show()

    def Delete(self):
        string = int(self.tableWidget.currentRow())
        id = str(self.tableWidget.item(string, 0).text())
        db.Delete(self.comboBox.currentText(), id)
        self.Show()


def main():
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle("Fusion")
    window = ExampleApp()
    window.show()
    app.exec_()


if __name__ == '__main__':

    main()
